from openpyxl import Workbook
wb = Workbook()

ws = wb.active
ws['A1'] = "身高"
ws['B1'] = "体重"
#身高
height = [180,175,170,165,160]
#体重
weight = [85,75,60,55,50]

for i in range(len(height)):
    ws.cell(row= 2+i,column=1,value=height[i])
    ws.cell(row= 2+i,column=2,value=weight[i])

wb.save("sample.xlsx")
